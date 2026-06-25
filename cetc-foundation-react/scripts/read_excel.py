import json, sys, os
try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl not installed")

excel_path = sys.argv[1] if len(sys.argv) > 1 else ''
if not excel_path or not os.path.exists(excel_path):
    sys.exit(f"Excel file not found: {excel_path}")

wb = openpyxl.load_workbook(excel_path)
ws = wb.active
courses = []
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    vals = [cell.value for cell in row]
    if vals[1]:
        courses.append({
            'sno': vals[0],
            'name': str(vals[1]),
            'sector': str(vals[2]) if vals[2] else '',
            'duration': str(vals[3]) if vals[3] else '',
            'level': str(vals[4]) if vals[4] else ''
        })

sys.stdout.reconfigure(encoding='utf-8')
print(json.dumps(courses, ensure_ascii=False))
